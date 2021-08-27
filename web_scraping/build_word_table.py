from get_word_info import get_word_info
from time import sleep
# input: text list with just word in xslx
# output: text list with word and its info in xsls


def get_file_name():
    from datetime import datetime
    now_dateNtime = str(datetime.now().strftime("%Y%m%d_%H%M"))
    file_name = now_dateNtime + "_word_list"+".xlsx"
    return file_name


def create_new_xlsx(file_name_to_create):
    # create a xlsx with sheet title by "Result"
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Result"
    wb.save(filename=file_name_to_create)


def create_new_xlsx_for_word_list():
    # create_new_xsxl_for_word_list
    # input: -
    # output: file_name in YYYYMMDD_HHMM_word_list.xlsx form
    try:
        file_name = get_file_name()
        create_new_xlsx(file_name)
        return file_name
    except:
        pass


def get_word_list(word_list_to_read):
    # input: word_list_to_read in xlsx, where col A as word list
    # output: word_list
    from openpyxl import load_workbook
    wb = load_workbook(word_list_to_read)
    sheet = wb.active
    word_list_raw = sheet["A"]
    word_list = []
    for data in word_list_raw:
        word_list.append(data.value)
    return word_list


def write_word_list(word_dict, count_row):
    from openpyxl import load_workbook
    wb = load_workbook(word_list_to_write)
    sheet = wb.active
    col_list = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    dict_key = ['word', 'eng_pr', 'ame_pr', 'tenses',
                'defination', 'ame_pr_url', 'eng_pr_url']
    for i in range(len(word_dict)):
        write_cell = col_list[i] + str(count_row)
        sheet[write_cell].value = word_dict[dict_key[i]]
    wb.save(filename=word_list_to_write)


def copy_word_list(word_list_to_read, word_list_to_write):
    # input: word_list_to_read in xlsx, where col A as word list
    # output: word_list_to_write in xlsx, where col A as word list, same with word_list_to_read
    word_list = get_word_list(word_list_to_read)
    write_word_list(word_list, word_list_to_write)


word_list_to_write = create_new_xlsx_for_word_list()
word_list = get_word_list("word_list_to_read.xlsx")
count_row = 1
for word in word_list:
    print(count_row, word)
    word_dict = get_word_info(word)
    write_word_list(word_dict, count_row)
    count_row = count_row + 1
    sleep(0.05)
